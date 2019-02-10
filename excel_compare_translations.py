#!python3
# expects excel file, translated or reviewed, with provided source and translation columns
# opens the file, extract data from provided columns, return dict 'translation_contents'
# dict structure: translation_contents[<file_name>][<tab_name>][<row_num>]{source, translation, reviewed_translation}

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def excel_contents(filepath, source_col, target_col, trans_or_rev):
    # open excel file
    try:
        wb = load_workbook(filepath)
    except Exception:
        return {}

    # get file name
    file_name = os.path.split(filepath)[1]

    # iterate every sheet (tab) from file
    worksheets = wb.worksheets

    # dictionary for source and target contents
    translation_contents = {}

    for ws in worksheets:
        # get sheet (tab) name
        tab_name = ws.title

        # parse source column and target column to get their contents
        row_num = 1

        # check if file name and tab name are in 'translation_contents' dict
        if file_name not in translation_contents:
            translation_contents[file_name] = {}
        if tab_name not in translation_contents[file_name]:
            translation_contents[file_name][tab_name] = {}

        while row_num <= ws.max_row:
            current_row = str(row_num)
            # get source and target cells values
            source_cell = ws[source_col + current_row].value
            target_cell = ws[target_col + current_row].value

            # check if row number exists in 'translation_contents[file_name][tab_name]' dict
            if row_num not in translation_contents[file_name][tab_name]:
                translation_contents[file_name][tab_name][row_num] = {}

            translation_contents[file_name][tab_name][row_num].update({'source': source_cell, trans_or_rev: target_cell})

            # add 1 to row_num to go to next row
            row_num += 1

    wb.close()

    return(translation_contents)


def generate_excel_report(filepath, translation_contents):
    # check if filepath is dir or file
    if os.path.isfile(filepath):
        dirname = os.path.dirname(filepath)
    else:
        dirname = filepath

    # removing last folder (expected: translation or review folder)
    dirname_split = dirname.split('\\')[:-1]
    # folder that we will save the report
    dir_destination = ''
    for elem in dirname_split:
        dir_destination += elem
        if dirname_split[-1] != elem:
            dir_destination += '\\'

    # create report file
    wb_report = Workbook()
    ws = wb_report.active
    ws.title = 'Report'

    # set names of columns
    ws['A1'] = 'Translated File'
    ws.column_dimensions['A'].width = 31.32
    ws['B1'] = 'Sheet Name'
    ws.column_dimensions['B'].width = 20.88
    ws['C1'] = 'Row'
    ws.column_dimensions['C'].width = 7.83
    ws['D1'] = 'Source'
    ws.column_dimensions['D'].width = 39.15
    ws['E1'] = 'Translation'
    ws.column_dimensions['E'].width = 39.15
    ws['F1'] = 'Reviewed Translation'
    ws.column_dimensions['F'].width = 39.15
    ws['G1'] = 'Reviewed File'
    ws.column_dimensions['G'].width = 31.32
    ws['H1'] = 'Updated?'
    ws.column_dimensions['H'].width = 10.44

    # get contents from translated (not reviewed) files and save them in newly created excel report
    sheet_row_num = 2
    for file_key in translation_contents:
        for sheet_key in translation_contents[file_key]:
            for row_key in translation_contents[file_key][sheet_key]:
                for cell_key in translation_contents[file_key][sheet_key][row_key]:
                    if 'translation_review' not in translation_contents[file_key][sheet_key][row_key]:
                        current_row = str(sheet_row_num)
                        cell_value = translation_contents[file_key][sheet_key][row_key][cell_key]
                        if cell_key == 'source':
                            ws['A' + current_row] = file_key
                            ws['B' + current_row] = sheet_key
                            ws['C' + current_row] = row_key
                            ws['D' + current_row] = cell_value
                            ws['D' + current_row].alignment = Alignment(wrap_text=True)
                        else:
                            ws['E' + current_row] = cell_value
                            ws['E' + current_row].alignment = Alignment(wrap_text=True)
                sheet_row_num += 1

    # get contents from reviewed files and compare it vs report contents and provide translation
    columns = list(ws.columns)
    source_column = columns[3]

    sheet_row_num = 2
    for file_key in translation_contents:
        for sheet_key in translation_contents[file_key]:
            for row_key in translation_contents[file_key][sheet_key]:
                for cell_key in translation_contents[file_key][sheet_key][row_key]:
                    if 'translation_review' in translation_contents[file_key][sheet_key][row_key]:
                        current_row = str(sheet_row_num)
                        cell_value = translation_contents[file_key][sheet_key][row_key][cell_key]
                        if cell_key == 'source':
                            for src in source_column:
                                if cell_value == src.value:
                                    src_row = str(src.row)
                                    review = translation_contents[file_key][sheet_key][row_key]['translation_review']
                                    ws['F' + src_row] = review
                                    ws['F' + src_row].alignment = Alignment(wrap_text=True)
                                    ws['G' + src_row] = file_key
                sheet_row_num += 1

    # compare translation and review, mark differences
    translation_column = columns[4]

    for cell in translation_column[1:]:
        current_row = str(cell.row)
        translation = cell.value
        review = ws['F' + current_row].value
        if translation != review:
            ws['H' + current_row] = 'Yes'
            ws['H' + current_row].fill = PatternFill(fgColor='FF0000', fill_type='solid')
            ws['F' + current_row].font = Font(color='FF0000')
        else:
            ws['H' + current_row] = 'No'

    # save the file
    wb_report.save(os.path.join(dir_destination, 'test.xlsx'))
